"""
Excel to TV Display - Backend API v2
Application pour afficher des plannings Excel sur TV avec gestion automatique des dates
"""

import os
import asyncio
import json
import re
from contextlib import asynccontextmanager
from pathlib import Path
from datetime import datetime, date, timedelta
from typing import Optional, Dict
from fastapi import FastAPI, WebSocket, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import aiofiles
import openpyxl
from openpyxl.utils import get_column_letter

# ============ CONFIGURATION ============
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)
ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv"}
REFRESH_INTERVAL = 15 * 60

# ============ VARIABLES GLOBALES ============
current_file: Optional[Path] = None
connected_clients = set()
file_observer: Optional[Observer] = None
sheets_cache: Dict = {}
sheets_cache_mtime = 0
file_modified_time = None

# ============ UTILITAIRES ============
def get_week_dates(week_num: int, year: int = None) -> list:
    """Retourne les dates (datetime.date) du lundi au dimanche pour une semaine donnée"""
    if year is None:
        year = datetime.now().year
    
    # Premier jour de l'année
    jan_4 = date(year, 1, 4)  # Le 4 janvier est toujours dans la semaine 1
    week_1_monday = jan_4 - timedelta(days=jan_4.weekday())
    
    # Calculer le lundi de la semaine demandée
    target_monday = week_1_monday + timedelta(weeks=week_num - 1)
    
    # Générer les 7 jours
    return [target_monday + timedelta(days=i) for i in range(7)]

def format_date_cell(day_date: date) -> str:
    """Formate une date pour l'affichage dans une cellule"""
    jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
    day_name = jours[day_date.weekday()]
    date_str = day_date.strftime("%d-%m-%Y")
    return f"{day_name}<br/>{date_str}"

def format_cell_value(value, row_idx: int = None, col_idx: int = None, 
                      week_dates: list = None, is_date_row: bool = False) -> str:
    """Formate une valeur de cellule avec reconstruction des dates si nécessaire"""
    try:
        # Ignorer les formules
        if isinstance(value, str) and value.startswith('='):
            # Si c'est une cellule de date dans la colonne A et qu'on a les dates de la semaine
            if is_date_row and col_idx == 1 and week_dates and row_idx:
                # Essayer de déterminer quel jour de la semaine c'est basé sur la position
                # (on suppose que les dates sont alignées avec les jours)
                day_index = (row_idx - 2) % 7  # Ajuster selon votre structure
                if 0 <= day_index < 7:
                    return format_date_cell(week_dates[day_index])
            return ""
        
        # Si c'est déjà une date/datetime
        if value and hasattr(value, 'strftime'):
            date_obj = value.date() if hasattr(value, 'date') else value
            return format_date_cell(date_obj)
        
        # Cellule vide dans une zone de dates : reconstruire si possible
        if (value is None or value == "") and is_date_row and col_idx == 1 and week_dates and row_idx:
            # Tenter de reconstruire la date manquante
            day_index = (row_idx - 2) % 7
            if 0 <= day_index < 7:
                return format_date_cell(week_dates[day_index])
        
        return str(value) if value is not None else ""
    except:
        return ""

def get_cell_style(cell) -> str:
    """Génère les styles CSS pour une cellule"""
    styles = []

    if cell.fill and cell.fill.start_color:
        try:
            rgb = cell.fill.start_color.rgb
            if rgb and rgb != "00000000":
                hex_color = str(rgb)[-6:] if len(str(rgb)) > 6 else str(rgb)
                if hex_color != "000000":
                    styles.append(f"background-color: #{hex_color}")
        except:
            pass

    if cell.font and cell.font.color:
        try:
            rgb = cell.font.color.rgb
            if rgb and rgb != "00000000":
                hex_color = str(rgb)[-6:] if len(str(rgb)) > 6 else str(rgb)
                if hex_color != "000000":
                    styles.append(f"color: #{hex_color}")
        except:
            pass
    
    if cell.font:
        if cell.font.bold:
            styles.append("font-weight: bold")
        if cell.font.italic:
            styles.append("font-style: italic")

    if cell.alignment:
        if cell.alignment.horizontal == "center":
            styles.append("text-align: center")
        if cell.alignment.vertical == "center":
            styles.append("vertical-align: middle")

    return "; ".join(styles)

def detect_date_rows(ws) -> set:
    """Détecte les lignes qui contiennent des dates dans la colonne A"""
    date_rows = set()
    for row_idx in range(1, min(30, ws.max_row + 1)):  # Scanner les 50 premières lignes
        cell = ws.cell(row_idx, 1)  # Colonne A
        if cell.value and hasattr(cell.value, 'strftime'):
            date_rows.add(row_idx)
        elif isinstance(cell.value, str) and cell.value.startswith('='):
            # Probablement une formule de date
            date_rows.add(row_idx)
    return date_rows

def sheet_to_html(ws, sheet_name: str = None) -> str:
    """Convertit une feuille Excel en HTML avec reconstruction des dates"""
    html = ['<table class="excel-table">']
    
    # Détecter le numéro de semaine à partir du nom de la feuille
    week_num = None
    week_dates = None
    if sheet_name and sheet_name.isdigit():
        week_num = int(sheet_name)
        week_dates = get_week_dates(week_num)
    
    # Détecter les lignes contenant des dates
    date_rows = detect_date_rows(ws)
    
    # Construire le dictionnaire des cellules fusionnées
    merged_cells_dict = {}
    for merged_range in ws.merged_cells.ranges:
        rowspan = merged_range.max_row - merged_range.min_row + 1
        colspan = merged_range.max_col - merged_range.min_col + 1
        
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                cell_coord = f"{get_column_letter(col)}{row}"
                is_main = (row == merged_range.min_row and col == merged_range.min_col)
                merged_cells_dict[cell_coord] = {
                    'rowspan': rowspan,
                    'colspan': colspan,
                    'is_main': is_main
                }
    
    # Parcourir toutes les lignes et colonnes
    current_day_index = 0  # Pour suivre quel jour on traite dans la colonne A
    max_column=20
    max_row=29
    #for row_idx in range(1, ws.max_row + 1):
    for row_idx in range(1, max_row + 1):
        html.append('  <tr>')
        
        #for col_idx in range(1, ws.max_column + 1):
        for col_idx in range(1, max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            
            # Ignorer les cellules fusionnées non-principales
            if cell.coordinate in merged_cells_dict:
                if not merged_cells_dict[cell.coordinate]['is_main']:
                    continue
            
            # Détecter si on est dans une zone de dates (colonne A)
            is_date_row = row_idx in date_rows or (col_idx == 1 and week_dates is not None)
            
            # Formater la valeur avec reconstruction des dates si nécessaire
            value = format_cell_value(
                cell.value, 
                row_idx=row_idx, 
                col_idx=col_idx,
                week_dates=week_dates,
                is_date_row=is_date_row
            )
            
            style = get_cell_style(cell)
            style_attr = f' style="{style}"' if style else ""
            
            # Ajouter les attributs de fusion si nécessaire
            if cell.coordinate in merged_cells_dict:
                merge_info = merged_cells_dict[cell.coordinate]
                if merge_info['colspan'] > 1:
                    style_attr += f' colspan="{merge_info["colspan"]}"'
                if merge_info['rowspan'] > 1:
                    style_attr += f' rowspan="{merge_info["rowspan"]}"'
            
            html.append(f'    <td{style_attr}>{value}</td>')
        
        html.append('  </tr>')
    
    html.append('</table>')
    return '\n'.join(html)

def filter_sheets(sheets: list) -> list:
    """Filtre les feuilles doublons"""
    return [s for s in sheets if not re.match(r'.*\s+\(\d+\)$', s)]

def convert_excel_file(file_path: Path) -> Dict:
    """Convertit un fichier Excel en HTML"""
    global sheets_cache, sheets_cache_mtime
    
    try:
        current_mtime = file_path.stat().st_mtime
        if sheets_cache and sheets_cache_mtime == current_mtime:
            return sheets_cache
        
        print(f"📂 Conversion: {file_path}")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets_html = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            html = sheet_to_html(ws, sheet_name)
            sheets_html[sheet_name] = html

        wb.close()
        sheets_cache = sheets_html
        sheets_cache_mtime = current_mtime
        print(f"✓ {len(sheets_html)} feuilles converties")
        return sheets_html

    except Exception as e:
        print(f"❌ Erreur: {e}")
        return {}

# ============ FILE WATCHER ============
class ExcelFileHandler(FileSystemEventHandler):
    def on_modified(self, event):
        if event.is_directory or not current_file:
            return
        if Path(event.src_path) == current_file:
            asyncio.create_task(notify_all_clients({
                "type": "file_modified",
                "timestamp": datetime.now().isoformat()
            }))

async def start_file_watcher():
    """Démarre la surveillance du fichier"""
    global file_observer
    if file_observer:
        file_observer.stop()
        file_observer.join()
    
    file_observer = Observer()
    file_observer.schedule(ExcelFileHandler(), str(UPLOAD_DIR), recursive=False)
    file_observer.start()
    print("✓ Surveillance activée")

# ============ WEBSOCKET ============
async def notify_all_clients(message: dict):
    """Notifie tous les clients"""
    disconnected = set()
    for client in connected_clients:
        try:
            await client.send_json(message)
        except:
            disconnected.add(client)
    connected_clients.difference_update(disconnected)

@asynccontextmanager
async def lifespan(app: FastAPI):
    global current_file, sheets_cache, sheets_cache_mtime, file_modified_time
    print("🚀 Serveur démarré")
    
    files = list(UPLOAD_DIR.glob("*.xlsx")) + list(UPLOAD_DIR.glob("*.xls"))
    if files:
        current_file = files[0]
        file_modified_time = datetime.fromtimestamp(current_file.stat().st_mtime)
        sheets_cache = {}
        sheets_cache_mtime = 0
        print(f"📂 Fichier auto-chargé: {current_file.name}")
        await start_file_watcher()
    
    yield
    
    global file_observer
    if file_observer:
        file_observer.stop()
        file_observer.join()
    print("🛑 Serveur arrêté")

# ============ FASTAPI APP ============
app = FastAPI(title="Excel to TV Display", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")

# ============ WEBSOCKET ============
@app.websocket("/ws")
async def websocket_endpoint(websocket):
    await websocket.accept()
    connected_clients.add(websocket)
    
    try:
        while True:
            data = await websocket.receive_text()
            msg = json.loads(data)
            if msg.get("type") == "ping":
                await websocket.send_json({"type": "pong"})
    except:
        pass
    finally:
        connected_clients.discard(websocket)

# ============ API ENDPOINTS ============
@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    """Upload un fichier Excel"""
    global current_file, sheets_cache, sheets_cache_mtime, file_modified_time
    
    if not any(file.filename.endswith(ext) for ext in ALLOWED_EXTENSIONS):
        raise HTTPException(status_code=400, detail="Format non supporté")
    
    file_path = UPLOAD_DIR / file.filename
    
    try:
        content = await file.read()
        async with aiofiles.open(file_path, "wb") as f:
            await f.write(content)
        
        current_file = file_path
        file_modified_time = datetime.now()
        sheets_cache = {}
        sheets_cache_mtime = 0
        
        await start_file_watcher()
        print(f"✓ Fichier uploadé: {file.filename}")
        
        return {"status": "success", "filename": file.filename}
    except Exception as e:
        print(f"❌ Erreur upload: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/sheets")
async def get_sheets():
    """Retourne les feuilles"""
    if not current_file or not current_file.exists():
        raise HTTPException(status_code=404, detail="Aucun fichier chargé")
    
    sheets_html = convert_excel_file(current_file)
    sheets = filter_sheets(list(sheets_html.keys()))
    return {"sheets": sheets}

@app.get("/sheet/{sheet_name}")
async def get_sheet(sheet_name: str):
    """Retourne une feuille en HTML"""
    if not current_file or not current_file.exists():
        raise HTTPException(status_code=404, detail="Aucun fichier chargé")
    
    sheets_html = convert_excel_file(current_file)
    if sheet_name not in sheets_html:
        raise HTTPException(status_code=404, detail="Feuille non trouvée")
    
    return {
        "sheet_name": sheet_name,
        "html": sheets_html[sheet_name],
        "timestamp": datetime.now().isoformat()
    }

@app.get("/files")
async def get_files():
    """Liste des fichiers"""
    files = [f.name for f in UPLOAD_DIR.glob("*.xlsx")]
    files.extend([f.name for f in UPLOAD_DIR.glob("*.xls")])
    return {"files": sorted(files)}

@app.get("/file-info")
async def get_file_info():
    """Info sur le fichier actuel"""
    if not current_file:
        return {"file": None, "modified": None}
    
    mod_time = datetime.fromtimestamp(current_file.stat().st_mtime)
    return {
        "file": current_file.name,
        "modified": mod_time.strftime("%d/%m/%Y %H:%M:%S")
    }

@app.get("/status")
async def get_status():
    """Statut du serveur"""
    return {
        "status": "running",
        "current_file": current_file.name if current_file else None,
        "connected_clients": len(connected_clients)
    }

# ============ PAGE PRINCIPALE ============
@app.get("/")
async def root():
    """Page HTML principale"""
    if not current_file or not current_file.exists():
        return HTMLResponse("""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Planning</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        html, body { width: 100%; height: 100%; }
        body { font-family: Arial, sans-serif; background: #1a1a1a; color: #fff; display: flex; justify-content: center; align-items: center; }
        .container { background: #2d2d2d; padding: 40px; border-radius: 8px; text-align: center; max-width: 500px; }
        h1 { font-size: 24px; margin-bottom: 20px; }
        .form-group { margin: 20px 0; }
        label { display: block; margin-bottom: 10px; font-weight: bold; }
        input, button { padding: 10px; font-size: 14px; border: 1px solid #0066cc; border-radius: 4px; }
        input[type="file"], input[type="text"] { width: 100%; background: #1a1a1a; color: #fff; margin-bottom: 10px; }
        button { background: #0066cc; color: white; cursor: pointer; font-weight: bold; }
        button:hover { background: #0052a3; }
        .divider { margin: 20px 0; color: #666; }
        .error { color: #ff6b6b; margin-top: 10px; display: none; }
        .success { color: #51cf66; margin-top: 10px; display: none; }
    </style>
</head>
<body>
    <div class="container">
        <h1>📊 Planning</h1>
        <p>Aucun fichier chargé</p>
        
        <div class="form-group">
            <label>Charger depuis URL:</label>
            <input type="text" id="urlInput" placeholder="https://exemple.com/planning.xlsx">
            <button onclick="loadFromUrl()">Charger depuis URL</button>
        </div>
        
        <div class="divider">OU</div>
        
        <div class="form-group">
            <label>Charger un fichier:</label>
            <input type="file" id="fileInput" accept=".xlsx,.xls,.csv">
            <button onclick="uploadFile()">Charger le fichier</button>
        </div>
        
        <div class="error" id="error"></div>
        <div class="success" id="success"></div>
    </div>
    
    <script>
        async function loadFromUrl() {
            const url = document.getElementById('urlInput').value;
            if (!url) { showError('Veuillez entrer une URL'); return; }
            try {
                showError('');
                const response = await fetch(url);
                if (!response.ok) throw new Error('Impossible de télécharger');
                const blob = await response.blob();
                const filename = url.split('/').pop() || 'planning.xlsx';
                await uploadFileContent(new File([blob], filename));
            } catch (e) {
                showError('Erreur: ' + e.message);
            }
        }
        
        async function uploadFile() {
            const file = document.getElementById('fileInput').files[0];
            if (!file) { showError('Sélectionnez un fichier'); return; }
            await uploadFileContent(file);
        }
        
        async function uploadFileContent(file) {
            try {
                const formData = new FormData();
                formData.append('file', file);
                const response = await fetch('/upload', { method: 'POST', body: formData });
                if (!response.ok) throw new Error('Erreur upload');
                showSuccess('Fichier chargé! Redirection...');
                setTimeout(() => window.location.reload(), 2000);
            } catch (e) {
                showError('Erreur: ' + e.message);
            }
        }
        
        function showError(msg) {
            const el = document.getElementById('error');
            el.textContent = msg;
            el.style.display = msg ? 'block' : 'none';
        }
        
        function showSuccess(msg) {
            const el = document.getElementById('success');
            el.textContent = msg;
            el.style.display = msg ? 'block' : 'none';
        }
    </script>
</body>
</html>""")
    
    # Page avec planning
    sheets_html = convert_excel_file(current_file)
    sheets = filter_sheets(list(sheets_html.keys()))
    
    today = date.today()
    week_num = today.isocalendar()[1]
    day_name = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"][today.weekday()]
    date_str = today.strftime("%d %b %Y")
    
    sheet_name = str(week_num) if str(week_num) in sheets else sheets[0]
    sheet_html = sheets_html[sheet_name]
    
    mod_time = datetime.fromtimestamp(current_file.stat().st_mtime).strftime("%d/%m/%Y %H:%M:%S")
    
    options_html = "\n".join([
        f'<option value="{sname}" {"selected" if sname == sheet_name else ""}>{sname}</option>'
        for sname in sheets
    ])
    
    html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Planning</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        html, body {{ width: 100%; height: 100%; }}
        body {{ font-family: Arial, sans-serif; background: #1a1a1a; color: #fff; overflow: hidden; display: flex; flex-direction: column; }}
        .header {{ background: #0a0a0a; padding: 10px 20px; border-bottom: 2px solid #0066cc; display: flex; justify-content: space-between; align-items: center; flex-shrink: 0; min-height: 45px; }}
        .header h1 {{ font-size: 18px; flex: 1; }}
        .day-date {{ font-size: 12px; font-weight: bold; color: #0099ff; }}
        .header-info {{ font-size: 11px; color: #999; text-align: right; }}
        .mod-time {{ font-size: 10px; color: #666; }}
        .controls {{ background: #2d2d2d; padding: 8px 20px; border-bottom: 1px solid #0066cc; display: flex; gap: 15px; align-items: center; flex-shrink: 0; height: 40px; }}
        .controls label {{ font-size: 13px; font-weight: bold; }}
        .controls button {{ background: #0066cc; color: white; border: none; border-radius: 4px; font-weight: bold; }}
        .controls button:hover {{ background: #0052a3; }}
        select {{ padding: 5px 10px; font-size: 13px; border: 1px solid #0066cc; border-radius: 4px; background: #1a1a1a; color: #fff; cursor: pointer; min-width: 80px; }}
        .content {{ flex: 1; background: #f5f5dc; overflow: auto; position: relative; }}
        .excel-table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
        .excel-table td {{ border: 1px solid #ddd; padding: 8px; text-align: left; white-space: pre-wrap; line-height: 1.3; background: #ffffff; }}
        .excel-table tr:nth-child(even) td {{ background: #fafaf5; }}
         table tr:first-child td {{font-size:1.3rem; font-weight: bold; text-align: center;}}
        /* Colonne A : fond bleu UNIQUEMENT si elle contient du texte */
        .excel-table td:first-child:not(:empty) {{ background-color: #4472C4 !important; color: #FFFFFF !important; font-weight: bold; text-align: center; }}
        ::-webkit-scrollbar {{ width: 8px; height: 8px; }}
        ::-webkit-scrollbar-track {{ background: #2d2d2d; }}
        ::-webkit-scrollbar-thumb {{ background: #0066cc; border-radius: 4px; }}
        @media (max-width: 1920px) {{
            .excel-table {{ font-size: clamp(12px, 1.2vw, 14px); }}
            .excel-table td {{ padding: clamp(6px, 0.8vw, 8px); }}
        }}

        @media (max-width: 1440px) {{
            .excel-table {{ font-size: clamp(11px, 1.1vw, 13px); }}
            .excel-table td {{ padding: clamp(5px, 0.7vw, 7px); }}
        }}

        @media (max-width: 1024px) {{
            .excel-table {{ font-size: clamp(10px, 1vw, 12px); }}
            .excel-table td {{ padding: clamp(4px, 0.6vw, 6px); }}
        }}

        @media (max-width: 768px) {{
            .excel-table {{ font-size: 10px; }}
            .excel-table td {{ padding: 4px; }}
            .excel-table {{
                display: block;
                overflow-x: auto;
            }}
        }}

    </style>
</head>
<body>
    <div class="header">
        <h1>📊 Planning</h1>
        <div class="header-info">
            <div class="day-date">{day_name} {date_str}</div>
            <div>Semaine {week_num} • {current_file.name}</div>
            <div class="mod-time">MAJ: {mod_time}</div>
        </div>
    </div>
    <div class="controls">
        <button onclick="navigateWeek(-1)" style="padding: 5px 15px; cursor: pointer;">◄ Précédent</button>
        <label>Semaine </label>
        <select id="sheet-select" onchange="changeSheet(this.value)">
            {options_html}
        </select>
        <button onclick="navigateWeek(1)" style="padding: 5px 15px; cursor: pointer;">Suivant ►</button>
        
        <span style="margin-left: auto; display: flex; gap: 10px; align-items: center;">
            <label>Zoom</label>
            <button onclick="adjustZoom(-10)">-</button>
            <span id="zoom-level">100%</span>
            <button onclick="adjustZoom(10)">+</button>
            <button onclick="resetZoom()">Reset</button>
        </span>
    </div>
    <div class="content" id="content">{sheet_html}</div>
    <script>
        let currentZoom = 100;
        
        function changeSheet(sname) {{
            fetch(`/sheet/${{encodeURIComponent(sname)}}`)
                .then(r => r.json())
                .then(d => {{
                    document.getElementById('content').innerHTML = d.html;
                    applyZoom();
                }})
                .catch(e => console.error(e));
        }}
        
        function navigateWeek(direction) {{
            const select = document.getElementById('sheet-select');
            const currentIndex = select.selectedIndex;
            const newIndex = currentIndex + direction;
            
            if (newIndex >= 0 && newIndex < select.options.length) {{
                select.selectedIndex = newIndex;
                changeSheet(select.value);
            }}
        }}
        
        function adjustZoom(delta) {{
            currentZoom = Math.max(50, Math.min(200, currentZoom + delta));
            applyZoom();
        }}
        
        function resetZoom() {{
            currentZoom = 100;
            applyZoom();
        }}
        
        function applyZoom() {{
            const content = document.getElementById('content');
            content.style.transform = `scale(${{currentZoom / 100}})`;
            content.style.transformOrigin = 'top left';
            content.style.width = `${{10000 / currentZoom}}%`;
            content.style.height = `${{10000 / currentZoom}}%`;
            document.getElementById('zoom-level').textContent = currentZoom + '%';
        }}
        
        setInterval(() => changeSheet(document.getElementById('sheet-select').value), {REFRESH_INTERVAL * 1000});
        
        // WebSocket pour les mises à jour en temps réel
        const protocol = window.location.protocol === 'https:' ? 'wss:' : 'ws:';
        const ws = new WebSocket(protocol + '//' + window.location.host + '/ws');
        ws.onmessage = (e) => {{
            const msg = JSON.parse(e.data);
            if (msg.type === 'file_modified') {{
                changeSheet(document.getElementById('sheet-select').value);
            }}
        }};
        ws.onopen = () => setInterval(() => ws.send(JSON.stringify({{type:'ping'}})), 30000);
    </script>
</body>
</html>"""
    return HTMLResponse(html)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)